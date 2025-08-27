import pandas as pd
from datetime import datetime
import os
from typing import Dict, Any
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

class ReportGenerator:
    def __init__(self, db_manager, analyzer):
        self.db_manager = db_manager
        self.analyzer = analyzer
        self.reports_dir = "/tmp/reports"
        os.makedirs(self.reports_dir, exist_ok=True)
    
    async def generate_excel_report(self, year: int, month: int) -> str:
        """Generate comprehensive Excel report for the month"""
        
        # Get monthly compliance report
        report_data = await self.analyzer.generate_monthly_compliance_report(year, month)
        
        # Get data quality issues
        data_issues = await self.analyzer.detect_data_quality_issues()
        
        # Get department statistics
        dept_stats = await self.analyzer.get_department_statistics(year, month)
        
        # Create Excel file
        filename = f"attendance_report_{year}_{month:02d}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Summary sheet
            await self._create_summary_sheet(writer, report_data, dept_stats)
            
            # Employee details sheet
            await self._create_employee_details_sheet(writer, report_data)
            
            # Data quality issues sheet
            await self._create_data_quality_sheet(writer, data_issues)
            
            # Department statistics sheet
            await self._create_department_sheet(writer, dept_stats)
        
        # Apply formatting
        self._format_excel_file(filepath)
        
        return filepath
    
    async def _create_summary_sheet(self, writer, report_data: Dict[str, Any], dept_stats):
        """Create summary sheet with key metrics"""
        
        summary_data = {
            'Metric': [
                'Total Employees',
                'Compliant Employees', 
                'Partially Compliant',
                'Non-Compliant',
                'Employees with Warnings',
                'Overall Compliance Rate (%)',
                'Month/Year'
            ],
            'Value': [
                report_data['total_employees'],
                report_data['compliance_summary']['compliant'],
                report_data['compliance_summary']['partial'],
                report_data['compliance_summary']['non_compliant'],
                report_data['compliance_summary']['warning'],
                round((report_data['compliance_summary']['compliant'] / 
                      report_data['total_employees']) * 100, 2) if report_data['total_employees'] > 0 else 0,
                f"{report_data['month']:02d}/{report_data['year']}"
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Add department summary
        if dept_stats:
            dept_summary = pd.DataFrame([
                {
                    'Department': dept.department_name,
                    'Total Employees': dept.total_employees,
                    'Compliant': dept.compliant_employees,
                    'Compliance Rate (%)': round(dept.average_compliance_rate, 2),
                    'Avg Hours/Employee': round(dept.average_hours_per_employee, 2),
                    'Data Issues': dept.total_data_issues
                }
                for dept in dept_stats
            ])
            
            # Add department summary starting from row 10
            dept_summary.to_excel(writer, sheet_name='Summary', startrow=10, index=False)
    
    async def _create_employee_details_sheet(self, writer, report_data: Dict[str, Any]):
        """Create detailed employee sheet"""
        
        employee_details = []
        
        for emp_stat in report_data['employee_stats']:
            employee = emp_stat['employee_info']
            stats = emp_stat['stats']
            
            employee_details.append({
                'Employee ID': employee['ID_PERSONA'],
                'Name': f"{employee['NOMBRE']} {employee['APELLIDO']}",
                'Department': employee.get('DEPARTAMENTO', 'N/A'),
                'Email': employee.get('EMAIL', 'N/A'),
                'Days Attended': stats.total_days_attended,
                'Total Hours': round(stats.total_hours_worked, 2),
                'Avg Hours/Day': round(stats.average_hours_per_day, 2),
                'Weeks with 1 Day': stats.weeks_with_1_day,
                'Weeks with 2 Days': stats.weeks_with_2_days,
                'Days Compliance': stats.days_compliance.value,
                'Hours Compliance': stats.hours_compliance.value,
                'Pattern Compliance': 'Yes' if stats.pattern_compliance else 'No',
                'Overall Compliance': stats.overall_compliance.value
            })
        
        employee_df = pd.DataFrame(employee_details)
        employee_df.to_excel(writer, sheet_name='Employee Details', index=False)
    
    async def _create_data_quality_sheet(self, writer, data_issues):
        """Create data quality issues sheet"""
        
        if not data_issues:
            # Create empty sheet with headers
            issues_df = pd.DataFrame(columns=[
                'Employee ID', 'Employee Name', 'Date', 'Issue Type', 
                'Description', 'Total Records', 'First Record', 'Last Record'
            ])
        else:
            issues_data = [
                {
                    'Employee ID': issue.employee_id,
                    'Employee Name': issue.employee_name,
                    'Date': issue.date.strftime('%Y-%m-%d'),
                    'Issue Type': issue.issue_type.value,
                    'Description': issue.description,
                    'Total Records': issue.total_records,
                    'First Record': issue.first_record.strftime('%Y-%m-%d %H:%M') if issue.first_record else 'N/A',
                    'Last Record': issue.last_record.strftime('%Y-%m-%d %H:%M') if issue.last_record else 'N/A'
                }
                for issue in data_issues
            ]
            issues_df = pd.DataFrame(issues_data)
        
        issues_df.to_excel(writer, sheet_name='Data Quality Issues', index=False)
    
    async def _create_department_sheet(self, writer, dept_stats):
        """Create department statistics sheet"""
        
        if not dept_stats:
            dept_df = pd.DataFrame(columns=[
                'Department', 'Total Employees', 'Compliant Employees',
                'Compliance Rate (%)', 'Avg Hours per Employee', 'Data Issues'
            ])
        else:
            dept_data = [
                {
                    'Department': dept.department_name,
                    'Total Employees': dept.total_employees,
                    'Compliant Employees': dept.compliant_employees,
                    'Compliance Rate (%)': round(dept.average_compliance_rate, 2),
                    'Avg Hours per Employee': round(dept.average_hours_per_employee, 2),
                    'Data Issues': dept.total_data_issues
                }
                for dept in dept_stats
            ]
            dept_df = pd.DataFrame(dept_data)
        
        dept_df.to_excel(writer, sheet_name='Department Statistics', index=False)
    
    def _format_excel_file(self, filepath: str):
        """Apply formatting to Excel file"""
        
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            wb = load_workbook(filepath)
            
            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            compliant_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            partial_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            non_compliant_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format headers
                for cell in ws[1]:
                    if cell.value:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Apply conditional formatting for compliance status
                if sheet_name == 'Employee Details':
                    for row in ws.iter_rows(min_row=2):
                        compliance_cell = None
                        for cell in row:
                            if ws.cell(1, cell.column).value == 'Overall Compliance':
                                compliance_cell = cell
                                break
                        
                        if compliance_cell and compliance_cell.value:
                            if compliance_cell.value == 'compliant':
                                for cell in row:
                                    cell.fill = compliant_fill
                            elif compliance_cell.value == 'partial':
                                for cell in row:
                                    cell.fill = partial_fill
                            elif compliance_cell.value == 'non_compliant':
                                for cell in row:
                                    cell.fill = non_compliant_fill
                        
                        # Apply borders
                        for cell in row:
                            cell.border = border
            
            wb.save(filepath)
            
        except Exception as e:
            logger.error(f"Error formatting Excel file: {e}")
            # File will still be usable without formatting
    
    async def generate_csv_report(self, year: int, month: int) -> str:
        """Generate CSV report for basic data export"""
        
        report_data = await self.analyzer.generate_monthly_compliance_report(year, month)
        
        employee_details = []
        
        for emp_stat in report_data['employee_stats']:
            employee = emp_stat['employee_info']
            stats = emp_stat['stats']
            
            employee_details.append({
                'employee_id': employee['ID_PERSONA'],
                'name': f"{employee['NOMBRE']} {employee['APELLIDO']}",
                'department': employee.get('DEPARTAMENTO', 'N/A'),
                'days_attended': stats.total_days_attended,
                'total_hours': round(stats.total_hours_worked, 2),
                'avg_hours_per_day': round(stats.average_hours_per_day, 2),
                'days_compliance': stats.days_compliance.value,
                'hours_compliance': stats.hours_compliance.value,
                'overall_compliance': stats.overall_compliance.value
            })
        
        # Create CSV file
        filename = f"attendance_report_{year}_{month:02d}.csv"
        filepath = os.path.join(self.reports_dir, filename)
        
        df = pd.DataFrame(employee_details)
        df.to_csv(filepath, index=False)
        
        return filepath